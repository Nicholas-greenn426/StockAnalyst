import bs4
import requests
import openpyxl
import datetime

'''
Template to set up new stock prices

EXPLAIN
REQUEST THE WEB PAGE AND CHECK STATUS TO MAKE SURE THAT IT IS AVIALABLE
PULLS THE CSS SELECTOR TO GATHER PRICE

#TICKER
TEMPRESVAR = requests.get('')
TEMPRESVAR.raise_for_status()

TEMPSOUP = bs4.BeautifulSoup(TEMPRESVAR.text,'html.parser')
TEMPELEY = soup.select('TEMPCSS')
print("TICKER Stock price: " + TEMPELEY[0].text)

'''



print("Hello Welcome to the money market")

#savethedate = input("What is todays date and est time")

tech = "Technology"
print(tech)
#AMD
amd = requests.get('https://www.google.com/finance/quote/AMD:NASDAQ?hl=en')
amd.raise_for_status()

soup = bs4.BeautifulSoup(amd.text,'html.parser')
elemamd = soup.select('#yDmH0d > c-wiz.zQTmif.SSPGKf.u5wqUe > div > div.e1AOyf > div > main > div.Gfxi4 > div.VfPpkd-WsjYwc.VfPpkd-WsjYwc-OWXEXe-INsAgc.KC1dQ.Usd1Ac.AaN0Dd.QZMA8b > c-wiz > div > div:nth-child(1) > div > div.rPF6Lc > div > div:nth-child(1) > div > span > div > div')
print("AMD Stock Price: " + elemamd[0].text)

#INTEL
intc = requests.get('https://www.google.com/finance/quote/INTC:NASDAQ?hl=en')
intc.raise_for_status()

intelligentsoup = bs4.BeautifulSoup(intc.text,'html.parser')
intelelem = intelligentsoup.select('#yDmH0d > c-wiz.zQTmif.SSPGKf.u5wqUe > div > div.e1AOyf > div > main > div.Gfxi4 > div.VfPpkd-WsjYwc.VfPpkd-WsjYwc-OWXEXe-INsAgc.KC1dQ.Usd1Ac.AaN0Dd.QZMA8b > c-wiz > div > div:nth-child(1) > div > div.rPF6Lc > div > div:nth-child(1) > div > span > div > div')
print("INTC Stock Price: " + intelelem[0].text)

#ADR (arm)
adr = requests.get('https://www.google.com/finance/quote/ARM:NASDAQ?hl=en')
adr.raise_for_status()

adrsoup = bs4.BeautifulSoup(adr.text,'html.parser')
adreley = adrsoup.select('#yDmH0d > c-wiz.zQTmif.SSPGKf.u5wqUe > div > div.e1AOyf > div > main > div.Gfxi4 > div.VfPpkd-WsjYwc.VfPpkd-WsjYwc-OWXEXe-INsAgc.KC1dQ.Usd1Ac.AaN0Dd.QZMA8b > c-wiz > div > div:nth-child(1) > div > div.rPF6Lc > div > div:nth-child(1) > div > span > div > div')
print("ARM Stock price: " + adreley[0].text)

#NVIDIA
nvidia = requests.get('https://www.google.com/finance/quote/NVDA:NASDAQ?hl=en')
nvidia.raise_for_status()

nvidiasoup = bs4.BeautifulSoup(nvidia.text,'html.parser')
nvidiaeley = nvidiasoup.select('#yDmH0d > c-wiz.zQTmif.SSPGKf.u5wqUe > div > div.e1AOyf > div > main > div.Gfxi4 > div.VfPpkd-WsjYwc.VfPpkd-WsjYwc-OWXEXe-INsAgc.KC1dQ.Usd1Ac.AaN0Dd.QZMA8b > c-wiz > div > div:nth-child(1) > div > div.rPF6Lc > div > div:nth-child(1) > div > span > div > div')
print("NVIDIA Stock price: " + nvidiaeley[0].text)

#print(savethedate)


#wb.get_sheet_names()


#letter = 'A'

#The following two variable can be used to create a new work book if nessicarry
#wb = openpyxl.Workbook()
#sheet = wb.get_sheet_by_name('Sheet')


workbook = openpyxl.load_workbook('Bigmoney.xlsx')
sheet1 = workbook.get_sheet_by_name('Sheet')

number = 2
colm = 1

while True:

    cell = sheet1.cell(row=number, column=colm).value 

    if cell is None or cell =='':
        break
    else:
        number += 1

sheet1['A1'] = 'DATE'
sheet1['C1'] = 'AMD'
sheet1['D1'] = 'INTEL'
sheet1['E1'] = 'ARM'
sheet1['F1'] = 'NVIDIA'



sheet1['A' + str(number)] = current_datetime = datetime.datetime.now()
sheet1['C' + str(number)] = elemamd[0].text
sheet1['D' + str(number)] = intelelem[0].text
sheet1['E' + str(number)] = adreley[0].text
sheet1['F' + str(number)] = nvidiaeley[0].text

workbook.save('Bigmoney.xlsx') 